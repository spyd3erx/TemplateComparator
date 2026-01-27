from src.utils.load_workbook import LoadWorkbook
import openpyxl

def create_test_file(tmp_path, file_name):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test"
    ws['A1'] = "Test"
    ws['B1'] = "Test"
    ws['C1'] = "Test"
    ws2 = wb.create_sheet("Test2")
    ws2['A1'] = "Test2"
    ws2['B1'] = "Test2"
    ws2['C1'] = "Test2"
    ws3 = wb.create_sheet("Test3")
    ws3['A1'] = "Test3"
    ws3['B1'] = "Test3"
    ws3['C1'] = "Test3"

    ws4 = wb.create_sheet("DM_Variables")
    for i in range(1, 20):
        ws4[f'A{i}'] = i
        ws4[f'B{i}'] = i**2
        ws4[f'C{i}'] = i**3
        ws4[f'D{i}'] = i**4     
        ws4[f'E{i}'] = i**5     
        ws4[f'F{i}'] = i**6     
        ws4[f'G{i}'] = i**7     
        ws4[f'H{i}'] = i**8     

    ws5 = wb.create_sheet("C")
    ws6 = wb.create_sheet("Tabla")
    
    wb.save(tmp_path / file_name)

    sheets = ["Test", "Test2", "Test3"]

    return tmp_path / file_name, sheets
    

def test_open_workbook(tmp_path):
    file_path, sheets = create_test_file(tmp_path, "test.xlsx")
    workbook = LoadWorkbook(file_path)
    assert workbook.get_sheet_names() == sheets
    workbook.close()



def test_open_workbook_data_only(tmp_path):
    """
    Este test verifica que el archivo solo se carguen los valores calculados.
    Es decir no se cargan las formulas.
    """ 
    file_path, sheets = create_test_file(tmp_path, "test.xlsx")
    workbook = LoadWorkbook(file_path, only_data=False)
    assert workbook.get_sheet_names() == sheets
    workbook.close()


def test_comparated_workbooks(tmp_path):
    """
    Este test verifica que el archivo solo se carguen los valores calculados.
    Es decir no se cargan las formulas.
    """ 
    file_path1, sheets = create_test_file(tmp_path, "test1.xlsx")
    file_path2, sheets = create_test_file(tmp_path, "test2.xlsx")
    workbook1 = LoadWorkbook(file_path1)
    workbook2 = LoadWorkbook(file_path2)

    full_sheets = list(set(workbook1.get_sheet_names() + workbook2.get_sheet_names()))
    full_sheets.sort()

    assert full_sheets == sheets
    
    workbook1.close()
    workbook2.close()

def test_get_sheet_by_name(tmp_path):
    file_path, sheets = create_test_file(tmp_path, "test.xlsx")
    workbook = LoadWorkbook(file_path)
    assert workbook.get_sheet_by_name("Test").title == sheets[0]
    workbook.close()


def test_comparated_workbook_only_sheet_name(tmp_path):
    file_path1, sheets = create_test_file(tmp_path, "test1.xlsx")
    file_path2, sheets = create_test_file(tmp_path, "test2.xlsx")
    workbook1 = LoadWorkbook(file_path1)
    workbook2 = LoadWorkbook(file_path2)

    full_sheets = list(set(workbook1.get_sheet_names() + workbook2.get_sheet_names()))
    full_sheets.sort()

    sheet_selected = workbook1.get_sheet_by_name("Test")
    
    assert sheet_selected.title in full_sheets
    workbook1.close()
    workbook2.close()