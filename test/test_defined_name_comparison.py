from src.core import DefinedNameComparison
from openpyxl import Workbook
from pathlib import Path
from openpyxl.workbook.defined_name import DefinedName
from src.core.utils.recursive_ordering_dict import sort_complete_dict

def create_excel_with_defined_names(path, names):
    wb = Workbook()
    ws = wb.active

    ws["A1"] = 100
    ws["A2"] = 200

    for name, formula in names.items():
        new_range = DefinedName(name, attr_text=formula)
        wb.defined_names.add(new_range)

    wb.save(path)
    wb.close()


class TestDefinedNameComparison:
    root_dir = Path(__file__).parent

    def test_get_defined_names(self, tmp_path):
        file1 = tmp_path / "file1.xlsx"
        file2 = tmp_path / "file2.xlsx"

        create_excel_with_defined_names(
            file1,
            {"TOTAL_SALES": "SUM(SalesData)", "AVERAGE_SALES": "AVERAGE(SalesData)"},
        )

        create_excel_with_defined_names(
            file2,
            {"TOTAL_SALES": "SUM(SalesData)", "AVERAGE_SALES": "AVERAGE(SalesData)"},
        )

        dnc = DefinedNameComparison(file1, file2)
        wb1 = dnc.open_workbook(file1).workbook
        defined_names = dnc.get_defined_names(wb1)

        expected_names = {
            "TOTAL_SALES": "SUM(SalesData)",
            "AVERAGE_SALES": "AVERAGE(SalesData)",
        }

        assert defined_names == expected_names

    def test_defined_names_diff(self, tmp_path):
        file1 = tmp_path / "file1.xlsx"
        file2 = tmp_path / "file2.xlsx"

        create_excel_with_defined_names(
            file1,
            {
                "TOTAL_SALES": "SUM(SalesData)",
                "AVERAGE_SALES": "AVERAGE(Sheet!$A$1:$A$2)",
                "MAX_SALES": "MAX(SalesData)",
            },
        )

        create_excel_with_defined_names(
            file2,
            {
                "Total_Sales": "SUM(SalesData)",
                "Average_Sales": "MEDIAN(SalesData)",
                "Min_Sales": "MIN(SalesData)",
            },
        )

        dnc = DefinedNameComparison(file1, file2)
        diff = dnc.defined_names_diff()

        expected_diff = {
            "coincidencias": [{"nombre": "TOTAL_SALES", "valor": "SUM(SalesData)"}],
            "diferencias": [
                {
                    "nombre": "AVERAGE_SALES",
                    "file1": "AVERAGE(Sheet!$A$1:$A$2)",
                    "file2": "MEDIAN(SalesData)",
                }
            ],
            "solo_en_archivo1": [{"nombre": "MAX_SALES", "valor": "MAX(SalesData)"}],
            "solo_en_archivo2": [{"nombre": "MIN_SALES", "valor": "MIN(SalesData)"}],
        }
        assert diff == expected_diff

    def test_defined_names_diff_no_defined_names(self, tmp_path):
        file1 = tmp_path / "file1.xlsx"
        file2 = tmp_path / "file2.xlsx"

        create_excel_with_defined_names(file1, {})
        create_excel_with_defined_names(file2, {})

        dnc = DefinedNameComparison(file1, file2)
        diff = dnc.defined_names_diff()

        expected_diff = {
            "coincidencias": [],
            "diferencias": [],
            "solo_en_archivo1": [],
            "solo_en_archivo2": [],
        }
        assert diff == expected_diff

    def test_defined_names_diff_one_file_empty(self, tmp_path):
        file1 = tmp_path / "file1.xlsx"
        file2 = tmp_path / "file2.xlsx"

        create_excel_with_defined_names(
            file1,
            {
                "TOTAL_SALES": "SUM(SalesData)",
                "AVERAGE_SALES": "AVERAGE(Sheet!$A$1:$A$2)",
            },
        )

        create_excel_with_defined_names(file2, {})

        dnc = DefinedNameComparison(file1, file2)
        diff = dnc.defined_names_diff()

        expected_diff = {
            "coincidencias": [],
            "diferencias": [],
            "solo_en_archivo1": [
                {"nombre": "AVERAGE_SALES", "valor": "AVERAGE(Sheet!$A$1:$A$2)"},
                {"nombre": "TOTAL_SALES", "valor": "SUM(SalesData)"},
            ],
            "solo_en_archivo2": [],
        }
        # diff_result = DeepDiff(expected_diff, diff, ignore_order=True)
        # assert not diff_result
        print(diff)
        assert sort_complete_dict(expected_diff) == diff

    def test_defined_names_diff_prd_files(self):
        file1 = self.root_dir / "RV_Patrimonio_SAP.xlsm"
        file2 = self.root_dir / "RV_Patrimonio_DIS.xlsx"

        dnc = DefinedNameComparison(file1, file2)
        
        wb2_names = dnc.get_defined_names(dnc.open_workbook(file2).workbook)
        nombres_con_problemas = {"ENTIDADES_ASOCIADAS", "AI", "CA"}

        expected = {
            "coincidencias": [
                {"nombre": k, "valor": v} 
                for k, v in wb2_names.items() 
                if k not in nombres_con_problemas
            ],
            "diferencias": [
                {
                    "nombre": "ENTIDADES_ASOCIADAS",
                    "RV_Patrimonio_SAP": "Tabla!$F$167:$F$685",
                    "RV_Patrimonio_DIS": "Tabla!$F$167:$F$684",
                }
            ],
            "solo_en_archivo1": [{"nombre": "AI", "valor": "Individual!$CB$201"}],
            "solo_en_archivo2": [{"nombre": "CA", "valor": "Individual!$BP$10"}],
        }

        get_diferences = dnc.defined_names_diff()

        assert sort_complete_dict(get_diferences) == sort_complete_dict(expected)