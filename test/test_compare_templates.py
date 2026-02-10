from src.core import TemplateComparison
from src.core.compare_templates import comparar_excels_md
from openpyxl import Workbook
import pytest


def create_excel_with_data(path, sheet_data):
    """
    Crea un archivo Excel con datos de prueba.

    Args:
        path: Ruta del archivo a crear
        sheet_data: Dict con nombres de hojas como keys y lista de celdas como values
                   Cada celda es un dict: {"cell": "A1", "value": valor, "formula": fórmula}
    """
    wb = Workbook()
    # Eliminar la hoja por defecto
    wb.remove(wb.active)

    for sheet_name, cells in sheet_data.items():
        ws = wb.create_sheet(sheet_name)
        for cell_data in cells:
            cell = ws[cell_data["cell"]]
            if "formula" in cell_data:
                cell.value = cell_data["formula"]
            else:
                cell.value = cell_data["value"]

    wb.save(path)
    wb.close()


class TestTemplateComparison:
    """Tests para la clase TemplateComparison."""

    def test_init(self, tmp_path):
        """Test de inicialización de la clase."""
        file1 = tmp_path / "file1.xlsx"
        file2 = tmp_path / "file2.xlsx"

        create_excel_with_data(file1, {"Sheet1": [{"cell": "A1", "value": 10}]})
        create_excel_with_data(file2, {"Sheet1": [{"cell": "A1", "value": 10}]})

        comp = TemplateComparison(file1, file2)

        assert comp.path1 == file1
        assert comp.path2 == file2

    def test_compare_values_identical_sheets(self, tmp_path):
        """Test comparación de valores con hojas idénticas."""
        file1 = tmp_path / "file1.xlsx"
        file2 = tmp_path / "file2.xlsx"
        output_md = tmp_path / "reporte.md"

        sheet_data = {
            "Sheet1": [
                {"cell": "A1", "value": 100},
                {"cell": "A2", "value": 200},
                {"cell": "B1", "value": "texto"},
            ]
        }

        create_excel_with_data(file1, sheet_data)
        create_excel_with_data(file2, sheet_data)

        comp = TemplateComparison(file1, file2)
        comp.compare(compare_values=True, output_md=str(output_md))

        assert output_md.exists()
        content = output_md.read_text(encoding="utf-8")
        assert "## Hoja: Sheet1" in content
        assert "Sin diferencias" in content

    def test_compare_values_with_differences(self, tmp_path):
        """Test comparación de valores con diferencias."""
        file1 = tmp_path / "file1.xlsx"
        file2 = tmp_path / "file2.xlsx"
        output_md = tmp_path / "reporte.md"

        sheet_data1 = {
            "Sheet1": [
                {"cell": "A1", "value": 100},
                {"cell": "A2", "value": 200},
            ]
        }

        sheet_data2 = {
            "Sheet1": [
                {"cell": "A1", "value": 150},  # Diferente
                {"cell": "A2", "value": 200},
            ]
        }

        create_excel_with_data(file1, sheet_data1)
        create_excel_with_data(file2, sheet_data2)

        comp = TemplateComparison(file1, file2)
        comp.compare(compare_values=True, output_md=str(output_md))

        content = output_md.read_text(encoding="utf-8")
        assert "A1" in content
        assert "100" in content
        assert "150" in content
        # A2 no debería aparecer en las diferencias
        assert "A2" not in content or "200" in content

    def test_compare_formulas_identical(self, tmp_path):
        """Test comparación de fórmulas idénticas."""
        file1 = tmp_path / "file1.xlsx"
        file2 = tmp_path / "file2.xlsx"
        output_md = tmp_path / "reporte.md"

        sheet_data = {
            "Sheet1": [
                {"cell": "A1", "formula": "=SUM(B1:B10)"},
                {"cell": "A2", "formula": "=AVERAGE(C1:C5)"},
            ]
        }

        create_excel_with_data(file1, sheet_data)
        create_excel_with_data(file2, sheet_data)

        comp = TemplateComparison(file1, file2)
        comp.compare(compare_values=False, output_md=str(output_md))

        content = output_md.read_text(encoding="utf-8")
        assert "Sin diferencias" in content

    def test_compare_formulas_with_differences(self, tmp_path):
        """Test comparación de fórmulas con diferencias."""
        file1 = tmp_path / "file1.xlsx"
        file2 = tmp_path / "file2.xlsx"
        output_md = tmp_path / "reporte.md"

        sheet_data1 = {
            "Sheet1": [
                {"cell": "A1", "formula": "=SUM(B1:B10)"},
                {"cell": "A2", "formula": "=AVERAGE(C1:C5)"},
            ]
        }

        sheet_data2 = {
            "Sheet1": [
                {"cell": "A1", "formula": "=SUM(B1:B20)"},  # Diferente
                {"cell": "A2", "formula": "=AVERAGE(C1:C5)"},
            ]
        }

        create_excel_with_data(file1, sheet_data1)
        create_excel_with_data(file2, sheet_data2)

        comp = TemplateComparison(file1, file2)
        comp.compare(compare_values=False, output_md=str(output_md))

        content = output_md.read_text(encoding="utf-8")
        assert "A1" in content
        assert "B1:B10" in content
        assert "B1:B20" in content

    def test_compare_formulas_normalizes_absolute_references(self, tmp_path):
        """Test que la normalización elimina referencias absolutas ($)."""
        file1 = tmp_path / "file1.xlsx"
        file2 = tmp_path / "file2.xlsx"
        output_md = tmp_path / "reporte.md"

        sheet_data1 = {
            "Sheet1": [
                {"cell": "A1", "formula": "=$A$1+$B$1"},
            ]
        }

        sheet_data2 = {
            "Sheet1": [
                {"cell": "A1", "formula": "=A1+B1"},  # Sin $
            ]
        }

        create_excel_with_data(file1, sheet_data1)
        create_excel_with_data(file2, sheet_data2)

        comp = TemplateComparison(file1, file2)
        comp.compare(compare_values=False, output_md=str(output_md))

        content = output_md.read_text(encoding="utf-8")
        # Deberían ser iguales después de normalizar
        assert "Sin diferencias" in content

    def test_compare_formulas_normalizes_spaces(self, tmp_path):
        """Test que la normalización elimina espacios."""
        file1 = tmp_path / "file1.xlsx"
        file2 = tmp_path / "file2.xlsx"
        output_md = tmp_path / "reporte.md"

        sheet_data1 = {
            "Sheet1": [
                {"cell": "A1", "formula": "= SUM ( A1 : A10 )"},
            ]
        }

        sheet_data2 = {
            "Sheet1": [
                {"cell": "A1", "formula": "=SUM(A1:A10)"},
            ]
        }

        create_excel_with_data(file1, sheet_data1)
        create_excel_with_data(file2, sheet_data2)

        comp = TemplateComparison(file1, file2)
        comp.compare(compare_values=False, output_md=str(output_md))

        content = output_md.read_text(encoding="utf-8")
        # Deberían ser iguales después de normalizar
        assert "Sin diferencias" in content

    def test_compare_specific_sheet(self, tmp_path):
        """Test comparación de una hoja específica."""
        file1 = tmp_path / "file1.xlsx"
        file2 = tmp_path / "file2.xlsx"
        output_md = tmp_path / "reporte.md"

        sheet_data1 = {
            "Sheet1": [{"cell": "A1", "value": 100}],
            "Sheet2": [{"cell": "A1", "value": 200}],
        }

        sheet_data2 = {
            "Sheet1": [{"cell": "A1", "value": 150}],  # Diferente
            "Sheet2": [{"cell": "A1", "value": 250}],  # Diferente
        }

        create_excel_with_data(file1, sheet_data1)
        create_excel_with_data(file2, sheet_data2)

        comp = TemplateComparison(file1, file2)
        comp.compare(compare_values=True, sheet="Sheet1", output_md=str(output_md))

        content = output_md.read_text(encoding="utf-8")
        assert "## Hoja: Sheet1" in content
        assert "## Hoja: Sheet2" not in content

    def test_compare_invalid_sheet_raises_error(self, tmp_path):
        """Test que comparar una hoja inexistente lanza error."""
        file1 = tmp_path / "file1.xlsx"
        file2 = tmp_path / "file2.xlsx"
        output_md = tmp_path / "reporte.md"

        sheet_data = {"Sheet1": [{"cell": "A1", "value": 100}]}

        create_excel_with_data(file1, sheet_data)
        create_excel_with_data(file2, sheet_data)

        comp = TemplateComparison(file1, file2)

        with pytest.raises(ValueError, match="no existe en ambos archivos"):
            comp.compare(
                compare_values=True, sheet="NoExiste", output_md=str(output_md)
            )

    def test_compare_excludes_excluded_sheets(self, tmp_path):
        """Test que las hojas en EXCLUDE_PARAMS son excluidas."""
        from src.config import EXCLUDE_PARAMS

        file1 = tmp_path / "file1.xlsx"
        file2 = tmp_path / "file2.xlsx"
        output_md = tmp_path / "reporte.md"

        # Crear hojas, incluyendo algunas excluidas
        sheet_data = {
            "Sheet1": [{"cell": "A1", "value": 100}],
            EXCLUDE_PARAMS[0]: [{"cell": "A1", "value": 999}],  # Esta debe ser excluida
        }

        create_excel_with_data(file1, sheet_data)
        create_excel_with_data(file2, sheet_data)

        comp = TemplateComparison(file1, file2)
        comp.compare(compare_values=True, output_md=str(output_md))

        content = output_md.read_text(encoding="utf-8")
        assert "## Hoja: Sheet1" in content
        assert f"## Hoja: {EXCLUDE_PARAMS[0]}" not in content

    def test_compare_formulas_skips_static_values(self, tmp_path):
        """Test que en modo fórmulas, las celdas sin fórmulas se saltan."""
        file1 = tmp_path / "file1.xlsx"
        file2 = tmp_path / "file2.xlsx"
        output_md = tmp_path / "reporte.md"

        sheet_data1 = {
            "Sheet1": [
                {"cell": "A1", "value": 100},  # Valor estático
                {"cell": "A2", "formula": "=B1+B2"},  # Fórmula
            ]
        }

        sheet_data2 = {
            "Sheet1": [
                {"cell": "A1", "value": 500},  # Diferente, pero valor estático
                {"cell": "A2", "formula": "=B1+B2"},  # Igual
            ]
        }

        create_excel_with_data(file1, sheet_data1)
        create_excel_with_data(file2, sheet_data2)

        comp = TemplateComparison(file1, file2)
        comp.compare(compare_values=False, output_md=str(output_md))

        content = output_md.read_text(encoding="utf-8")
        # A1 no debe aparecer porque ambas son valores estáticos
        assert "Sin diferencias" in content

    def test_compare_multiple_sheets(self, tmp_path):
        """Test comparación de múltiples hojas."""
        file1 = tmp_path / "file1.xlsx"
        file2 = tmp_path / "file2.xlsx"
        output_md = tmp_path / "reporte.md"

        sheet_data = {
            "Sheet1": [{"cell": "A1", "value": 100}],
            "Sheet2": [{"cell": "A1", "value": 200}],
            "Sheet3": [{"cell": "A1", "value": 300}],
        }

        create_excel_with_data(file1, sheet_data)
        create_excel_with_data(file2, sheet_data)

        comp = TemplateComparison(file1, file2)
        comp.compare(compare_values=True, output_md=str(output_md))

        content = output_md.read_text(encoding="utf-8")
        assert "## Hoja: Sheet1" in content
        assert "## Hoja: Sheet2" in content
        assert "## Hoja: Sheet3" in content

    def test_compare_no_common_sheets(self, tmp_path):
        """Test comparación sin hojas comunes."""
        file1 = tmp_path / "file1.xlsx"
        file2 = tmp_path / "file2.xlsx"
        output_md = tmp_path / "reporte.md"

        sheet_data1 = {"Sheet1": [{"cell": "A1", "value": 100}]}
        sheet_data2 = {"Sheet2": [{"cell": "A1", "value": 200}]}

        create_excel_with_data(file1, sheet_data1)
        create_excel_with_data(file2, sheet_data2)

        comp = TemplateComparison(file1, file2)
        comp.compare(compare_values=True, output_md=str(output_md))

        # No debe generar archivo o debe estar vacío
        # La función retorna sin hacer nada si no hay hojas comunes
        # Verificamos que no haya contenido de comparación
        if output_md.exists():
            content = output_md.read_text(encoding="utf-8")
            assert content.strip() == ""

    def test_compare_handles_none_values(self, tmp_path):
        """Test que maneja correctamente valores None."""
        file1 = tmp_path / "file1.xlsx"
        file2 = tmp_path / "file2.xlsx"
        output_md = tmp_path / "reporte.md"

        sheet_data1 = {
            "Sheet1": [
                {"cell": "A1", "value": 100},
                {"cell": "A2", "value": None},
            ]
        }

        sheet_data2 = {
            "Sheet1": [
                {"cell": "A1", "value": 100},
                {"cell": "A2", "value": 0},  # None se normaliza a 0
            ]
        }

        create_excel_with_data(file1, sheet_data1)
        create_excel_with_data(file2, sheet_data2)

        comp = TemplateComparison(file1, file2)
        comp.compare(compare_values=True, output_md=str(output_md))

        content = output_md.read_text(encoding="utf-8")
        # None y 0 deberían ser iguales después de normalizar
        assert "Sin diferencias" in content


class TestComparationLegacyFunction:
    """Tests para la función legacy comparar_excels_md."""

    def test_legacy_function_compatibility(self, tmp_path):
        """Test que la función legacy funciona correctamente."""
        file1 = tmp_path / "file1.xlsx"
        file2 = tmp_path / "file2.xlsx"
        output_md = tmp_path / "reporte.md"

        sheet_data = {"Sheet1": [{"cell": "A1", "value": 100}]}

        create_excel_with_data(file1, sheet_data)
        create_excel_with_data(file2, sheet_data)

        comparar_excels_md(file1, file2, compare_values=True, salida_md=str(output_md))

        assert output_md.exists()
        content = output_md.read_text(encoding="utf-8")
        assert "## Hoja: Sheet1" in content

    def test_legacy_function_with_differences(self, tmp_path):
        """Test función legacy con diferencias."""
        file1 = tmp_path / "file1.xlsx"
        file2 = tmp_path / "file2.xlsx"
        output_md = tmp_path / "reporte.md"

        sheet_data1 = {"Sheet1": [{"cell": "A1", "value": 100}]}
        sheet_data2 = {"Sheet1": [{"cell": "A1", "value": 200}]}

        create_excel_with_data(file1, sheet_data1)
        create_excel_with_data(file2, sheet_data2)

        comparar_excels_md(file1, file2, compare_values=True, salida_md=str(output_md))

        content = output_md.read_text(encoding="utf-8")
        assert "A1" in content
        assert "100" in content
        assert "200" in content

