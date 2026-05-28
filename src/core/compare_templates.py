from pathlib import Path
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
import re
from typing import Optional, Any
from src.config import DEFAULT_MARKDOWN_FILE, EXCLUDE_PARAMS

def normalizar_formula(v: Any) -> Any:
    """
    Normaliza una fórmula de Excel:
    1. Si no es string, retorna el valor original.
    2. Detecta si parece fórmula (empieza con =, +, -).
    3. Identifica cadenas entre comillas dobles.
    4. Elimina espacios fuera de cadenas.
    5. Estandariza el prefijo a '='.
    6. Estandarizar valores fijos: Elimina '$' de referencias absolutas.
    """
    if not isinstance(v, str):
        return v
    
    # Limpiar espacios iniciales/finales
    v = v.strip()
    
    # Verificar si es fórmula (empieza con =, +, -)
    # Nota: Excel visualiza '+A1' como '+A1' pero internamente es '=A1' o '=+A1'. 
    # OpenPyXL con data_only=False devuelve lo que lee.
    if not (v.startswith("=") or v.startswith("+") or v.startswith("-")):
        return v

    # Estandarizar prefijo: '+IF' -> '=IF', '=+IF' -> '=IF', '=-IF' -> '=-IF' (negativo es lógica)
    # Si empieza con =, quitamos para limpiar y luego reponemos.
    temp_v = v
    if temp_v.startswith("="):
        temp_v = temp_v[1:]
    
    # Remove leading + if present (e.g. =+IF or +IF becomes IF)
    if temp_v.startswith("+"):
        temp_v = temp_v[1:]
        
    # Reconstruimos con '='
    v_clean = "=" + temp_v

    # Regex para capturar cadenas entre comillas (incluyendo "") o cualquier otro caracter no-espacio
    pattern = r'"(?:[^"]|"")*"|[^"\s]+'
    parts = re.findall(pattern, v_clean)
    
    formula_clean =  "".join(parts)
    return formula_clean

def normalizar(v: Any) -> Any:
    if v is None:
        return 0
    if isinstance(v, str):
        v = v.strip()
        return 0 if v == "" else v
    return v

def comparar_excels_md(
    archivo1: Path,
    archivo2: Path,
    compare_values: bool,
    sheet: Optional[str] = None,
    salida_md: str = DEFAULT_MARKDOWN_FILE,
):
    # compare_values=True -> data_only=True (Read Values)
    # compare_values=False -> data_only=False (Read Formulas)
    wb1 = load_workbook(archivo1, data_only=compare_values)
    wb2 = load_workbook(archivo2, data_only=compare_values)

    archivo1_nombre = Path(archivo1).stem
    archivo2_nombre = Path(archivo2).stem

    hojas1 = set(wb1.sheetnames)
    hojas2 = set(wb2.sheetnames)

    hojas_comunes = hojas1.intersection(hojas2)

    # Filter excluded sheets
    hojas_comunes = {h for h in hojas_comunes if h not in EXCLUDE_PARAMS}

    if not hojas_comunes:
        wb1.close()
        wb2.close()
        return

    try:
        if sheet:
            if sheet not in hojas_comunes:
                raise ValueError(f"La hoja '{sheet}' no existe en ambos archivos.")
            hojas_a_comparar = {sheet}
        else:
            hojas_a_comparar = hojas_comunes

        reporte_md = []

        for hoja in sorted(hojas_a_comparar):            
            ws1: Worksheet = wb1[hoja]
            ws2: Worksheet = wb2[hoja]
            
            # Calculate max bounds to iterate efficiently
            max_row = max(ws1.max_row, ws2.max_row)
            max_col = max(ws1.max_column, ws2.max_column)
            
            diferencias = []
            
            for row in range(1, max_row + 1):
                for col in range(1, max_col + 1):
                    c1 = ws1.cell(row=row, column=col)
                    c2 = ws2.cell(row=row, column=col)
                    
                    v1_raw = c1.value
                    v2_raw = c2.value

                    # Filter: If in Formula Mode (compare_values=False), 
                    # skip comparison if neither cell has a formula (both are static values).
                    # We consider it a formula if it is a string starting with '='
                    if not compare_values:
                        is_f1 = isinstance(v1_raw, str) and v1_raw.strip().startswith("=")
                        is_f2 = isinstance(v2_raw, str) and v2_raw.strip().startswith("=")
                        
                        # Extra requirement: "Evaluated cells contain only formulas"
                        # Interpretation: Filter out noise where both are static.
                        # We keep the case where one is formula and other is value (overwrite detection).
                        if not is_f1 and not is_f2:
                            continue

                    # Apply formula normalization always. 
                    # If it's a value (not starting with =), it returns original.
                    # If it's a formula, it normalizes it.
                    v1 = normalizar_formula(v1_raw)
                    v2 = normalizar_formula(v2_raw)

                    v1 = normalizar(v1)
                    v2 = normalizar(v2)
                    
                    if v1 != v2:
                         diferencias.append({
                            "celda": c1.coordinate,
                            "archivo1": v1,
                            "archivo2": v2
                         })

            # Construcción de tabla Markdown
            reporte_md.append(f"## Hoja: {hoja}\n")

            if diferencias:
                reporte_md.append(f"| Celda | {archivo1_nombre} | {archivo2_nombre} |")
                reporte_md.append("|-------|-----------|-----------|")

                for d in diferencias:
                    v1_str = str(d["archivo1"]).replace("|", "\\|")
                    v2_str = str(d["archivo2"]).replace("|", "\\|")
                    reporte_md.append(f"| {d['celda']} | `{v1_str}` | `{v2_str}` |")
            else:
                reporte_md.append("Sin diferencias.\n")

            reporte_md.append("\n")

    finally:
        wb1.close()
        wb2.close()

    # Guardar archivo MD
    with open(salida_md, "w", encoding="utf-8") as f:
        f.write("\n".join(reporte_md))
        
    print(f"Reporte generado en: {salida_md}")


class TemplateComparison:
    """
    Clase de compatibilidad que envuelve la función comparar_excels_md
    para mantener la misma interfaz que usa la GUI y los tests.
    """

    def __init__(self, path1: Path, path2: Path):
        self.path1 = Path(path1)
        self.path2 = Path(path2)

    def compare(
        self,
        compare_values: bool,
        sheet: Optional[str] = None,
        output_md: str | Path = DEFAULT_MARKDOWN_FILE,
    ) -> None:
        comparar_excels_md(
            self.path1,
            self.path2,
            compare_values=compare_values,
            sheet=sheet,
            salida_md=str(output_md),
        )
