from openpyxl import load_workbook
from pathlib import Path
from src.config import EXCLUDE_PARAMS


class LoadWorkbook:
    """
    Clase para cargar un archivo de excel.

    Args:
        file_path (str): Ruta del archivo de excel.

        only_read (bool): Indica si se debe cargar el archivo en modo de solo lectura. Default: True
        cuando se carga en este modo es necesario que se cierre el archivo explicitamente.

        only_data (bool): Indica si se debe cargar el archivo en modo de solo datos. Default: True
    """

    _excluded_sheets = set(EXCLUDE_PARAMS)

    def __init__(self, file_path, only_read=True, only_data=True):
        """
        Inicializa la clase LoadWorkbook.
        """
        file_path = Path(file_path)
        if not file_path.exists():
            raise FileNotFoundError(f"El archivo {file_path} no existe.")
        self.file_path = file_path
        self.workbook = load_workbook(
            file_path, read_only=only_read, data_only=only_data
        )

    def get_sheet_names(self) -> list[str]:
        """
        Retorna los nombres de las hojas filtrando las excluidas.
        """
        return [
            name
            for name in self.workbook.sheetnames
            if name not in self._excluded_sheets
        ]

    def get_sheet_by_name(self, sheet_name: str):
        """
        Obtiene una hoja del archivo de excel.

        Args:
            sheet_name (str): Nombre de la hoja.
        """
        return self.workbook[sheet_name]

    def close(self) -> None:
        """
        Cierra el archivo de excel.
        """
        self.workbook.close()

    def __enter__(self):
        return self

    def __exit__(self, *args):
        self.close()
